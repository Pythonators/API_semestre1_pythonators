from tkinter import *
from tkinter.ttk import *

import openpyxl
from openpyxl import Workbook, load_workbook
import xlsxwriter
from openpyxl import *
import numpy as np
import pandas as pd
from PIL import ImageTk, Image
from openpyxl.reader.excel import load_workbook
def AbrirJanelaAvaliacao():
    janela = Tk()
    largura_screen = janela.winfo_screenwidth()
    altura_screen = janela.winfo_screenheight()
    largura = 480
    altura = 800
    posx = largura_screen / 2 - largura / 2
    posy = (altura_screen / 2-30) - altura / 2
    janela.geometry('%dx%d+%d+%d' % (largura, altura, posx, posy))

    def avaliar():
        cred1 = int(pri.get())
        cred2 = int(seg.get())
        cred3 = int(tri.get())
        cred4 = int(quar.get())
        cred5 = int(quin.get())
        credenciais = []
        credenciais.append(cred1)
        credenciais.append(cred2)
        credenciais.append(cred3)
        credenciais.append(cred4)
        credenciais.append(cred5)

        wb = Workbook()
        sh = wb.active
        i = 0

        array = ['A','B','C','D','E']
        tam = len(array)
        while i < tam:
            sh[array[i]+(str(1))] = "PGT"+str(i+1)
            i+=1
        sh['A'+str(2)] = credenciais[0]
        sh['B'+str(2)] = credenciais[1]
        sh['C'+str(2)] = credenciais[2]
        sh['D'+str(2)] = credenciais[3]
        sh['E'+str(2)] = credenciais[4]
        i+=1

        wb.save(filename='data_base/Pasta1.xlsx')
        janela.destroy()

    main_frame = Frame(janela)
    main_frame.pack(fill=BOTH, expand=1)

    my_canvas = Canvas(main_frame, bg='red')
    my_canvas.pack(side=LEFT, fill=BOTH, expand=1)

    my_scrollbar = Scrollbar(main_frame, orient=VERTICAL, command=my_canvas.yview)
    my_scrollbar.pack(side=RIGHT, fill=Y)

    my_canvas.configure(yscrollcommand=my_scrollbar.set)
    my_canvas.bind('<Configure>', lambda e: my_canvas.configure(scrollregion= my_canvas.bbox("all")))

    second_frame = Frame(my_canvas)


    my_canvas.create_window((0,0), window=second_frame, anchor=NW)

    style = Style(janela)
    style.configure("TRadiobutton",font='Arial 16')
    canvas = Canvas(janela)
    scrolly = Scrollbar(janela, orient='vertical', command=canvas.yview)
    #Variaveis---------------------------------------------------
    pri = IntVar()
    seg = IntVar()
    tri = IntVar()
    quar = IntVar()
    quin = IntVar()

    MODES = [('Extremamente ', 1),
            ('Muito', 2),
            ('Razoável', 3),
            ('Pouco', 4),
             ('Nada',5)]


    #Perguntas----------------------------------------------------
    l_text= Label(second_frame, justify='left', font='Ivy 18', wraplength=450,
                    text='1 - O avaliado fez entregas pontualmente')
    l_text.pack( anchor=NW, padx=15, pady=5)

    for text,mode in MODES:
        Radiobutton(second_frame, text=text, variable=pri, value=mode).pack( anchor=NW, padx=15, pady=5)

    l_text1= Label(second_frame, justify='left', font='Ivy 18', wraplength=450,
                    text='2 - O avaliado fez entregas de acordo com as propostas da sprint')
    l_text1.pack( anchor=NW, padx=15, pady=5)

    for text,mode in MODES:
        Radiobutton(second_frame, text=text, variable=seg, value=mode).pack( anchor=NW, padx=15, pady=5)

    l_text2= Label(second_frame, justify='left', font='Ivy 18', wraplength=450,
                    text='3 - O avaliado teve um bom desempenho no trabalho em equipe')
    l_text2.pack( anchor=NW, padx=15, pady=5)

    for text,mode in MODES:
        Radiobutton(second_frame, text=text, variable=tri, value=mode).pack( anchor=NW, padx=15, pady=5)

    l_text3= Label(second_frame, justify='left', font='Ivy 18', wraplength=450,
                    text='4 - O avaliado demonstrou habilidades e/ou desejo em se desenvolver nas tecnologias usadas no projeto')
    l_text3.pack( anchor=NW, padx=15, pady=5)

    for text,mode in MODES:
        Radiobutton(second_frame, text=text, variable=quar, value=mode).pack( anchor=NW, padx=15, pady=5)

    l_text4= Label(second_frame, justify='left', font='Ivy 18', wraplength=450,
                    text='5 - O avaliado teve uma comunicação clara com o grupo quanto às suas dificuldades e evoluções no decorrer das sprints')
    l_text4.pack( anchor=NW, padx=15, pady=5)


    for text,mode in MODES:
        Radiobutton(second_frame, text=text, variable=quin, value=mode).pack( anchor=NW, padx=15, pady=5)

    i_peso = Image.open('Imagens/botao.png')
    i_peso = ImageTk.PhotoImage(i_peso)
    but = Button(second_frame,  image=i_peso, width=100, command= avaliar)
    but.pack(anchor=NW, padx=15, pady=5)
    print(MODES[1])

    janela.resizable(False,False)
    janela.mainloop()
