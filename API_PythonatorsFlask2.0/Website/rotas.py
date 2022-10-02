from flask import Flask, render_template, request, Blueprint
import openpyxl
from openpyxl import Workbook, load_workbook
import xlsxwriter
from openpyxl import *
import numpy as np
import pandas as pd
from openpyxl.reader.excel import load_workbook
from werkzeug import debug

app = Flask(__name__)
# app = Blueprint('app', __name__)
class Aluno:
    def __init__(self, nome):
        self.nome = nome


a1 = Aluno('Caio')
a2 = Aluno('Marcelo')
a4 = Aluno('Gabriel')
a5 = Aluno('Mikaéla')
a6 = Aluno('Isadora')
a7 = Aluno('Cauana')
a8 = Aluno('Ana')
a9 = Aluno('Felipe')
a10 = Aluno('Murilo')
a11 = Aluno('Ryan')

alunos = []
alunos.append(a1), alunos.append(a2), alunos.append(a4), alunos.append(a5), alunos.append(a6), alunos.append(a7),
alunos.append(a8), alunos.append(a9), alunos.append(a10), alunos.append(a11)
class Pergunta:
    def __init__(self, pergunta_, name):
        self.pergunta_ = pergunta_
        self.name = name


p1 = Pergunta('1 - O avaliado fez entregas pontualmente', '1')
p2 = Pergunta('2 - O avaliado fez entregas de acordo com as propostas da sprint', '2')
p3 = Pergunta('3 - O avaliado teve um bom desempenho no trabalho em equipe', '3')
p4 = Pergunta('4 - O avaliado demonstrou habilidades e/ou desejo em se desenvolver nas tecnologias usadas no projeto',
              '4')
perguntas = []
perguntas.append(p1)
perguntas.append(p2)
perguntas.append(p3)
perguntas.append(p4)
@app.route('/')
def pagina_login():
    return render_template("index.html")

@app.route('/sprint')
def pagina_sprint():
    return render_template("sprint.html")

@app.route('/admin')
def pagina_admin():
    return "<h1>página cadastro de alunos e grupos</h1>"

@app.route('/admin/cadastro')
def pagina_cadastro_admin():
    return render_template("admin_cadastro.html",textadmin = 'Hello, Admin')

@app.route("/aluno/avaliacao",)
def avaliacao():
    return render_template("avaliacao.html", alunos = alunos, perguntas = perguntas)


@app.route("/aluno/notas",methods=['POST'])
def aluno_notas():
        x = 0
        y = 0
        wb = Workbook()
        sh = wb.active
        array = ['A', 'B', 'C', 'D']
        array2 = ['A','B','C','D','E']

        option = []
        lista = [1,2,3,4]
        for x in range(len(alunos)):
            option.append([str(alunos[x].nome)])

        result = len(alunos)
        for x in range(result):
            for y in range(len(lista)):
                option[x].append(request.form[(str(alunos[x].nome+str(lista[y])))])
                y+=1
            x += 1
        tam = len(array)
        i = 0
        d = 0
        l = 1
        while i < tam:
            sh[array[i] + (str(1))] = "PGT" + str(i + 1)
            i += 1
        sh['E1'] = 'AVALIADO'

        for d in range(len(option)):


                sh['A' + str(l+1)] = option[d][1]
                sh['B' + str(l+1)] = option[d][2]
                sh['C' + str(l+1)] = option[d][3]
                sh['D' + str(l+1)] = option[d][4]
                sh['E' + str(l+1)] = option[d][0]
                l += 1

                d+=1


        i += 1

        wb.save(filename='templates/Pasta1.xlsx')
        # for z in range(len(option)):
        #     for x in range(len(alunos)):
        #         if option[z] == str(alunos[x].nome):
        #             option[z].append('{')
        #             break
        #         x+=1
        #     z+=1



        #
        #         # option = request.form[str(alunos[x].nome+perguntas[y].name)]
        #         x+=1
        # print(option)


        return option
    # # sh['A' + str(2)] = option
    # print(option)


@app.route('/professor-m2')
def professorm2():
    return "<h1>página professor M2</h1>"

@app.route('/professor-p2')
def professorp2():
    return "<h1>página professor P2</h1>"

@app.route("/professor/dashboard")
def professor_dash():
    return "<h1>página dashboard</h1>"



